from PyQt6.QtWidgets import (
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QStackedWidget,
    QMessageBox,
    QLabel,
    QButtonGroup,
    QFrame,
    QTabWidget,
    QToolButton,
    QStyle,
    QComboBox,
    QMenu,
    QDialog,
    QTextEdit,
    QDialogButtonBox,
    QCheckBox,
    QScrollArea,
    QSizePolicy,
)
from PyQt6.QtGui import QIcon, QAction, QPixmap, QPainter, QGuiApplication
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSize, QByteArray, QPropertyAnimation, QEasingCurve, QRect, QTimer
from PyQt6.QtSvg import QSvgRenderer
from pathlib import Path
from ui.file_drop import FileDropWidget
from ui.excel.result_view import ResultView
import traceback
import sys
import os
import re
import shutil
import tempfile
import subprocess
from openpyxl import load_workbook
from core.config import (
    get_last_open_dir,
    update_last_open_dir,
    get_update_prompt_on_startup,
    set_update_prompt_on_startup,
    set_pending_update,
    pop_pending_update,
)
from core.version import APP_VERSION, CURRENT_VERSION
from core.update_manager import check_for_update
from core.feedback_manager import submit_feedback
from core.utils import convert_xls_to_xlsx

# Add root to sys.path to ensure core imports work
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from excel.engine import load_excel_with_styles, compare_dataframes


class ClickableLabel(QLabel):
    clicked = pyqtSignal()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)


class UpdateCheckThread(QThread):
    finished = pyqtSignal(object, object)

    def __init__(self, current_version):
        super().__init__()
        self.current_version = current_version

    def run(self):
        info, err = check_for_update(self.current_version)
        self.finished.emit(info, err)

class WorkerThread(QThread):
    finished = pyqtSignal(object, object, object, object, object, object, object, object, object, object, object, object, object) 
    # df1, df2, styles1, styles2, merges1, merges2, formulas1, formulas2, col_widths1, row_heights1, col_widths2, row_heights2, result
    error = pyqtSignal(str)

    def __init__(self, file1, file2, sheet1_name, sheet2_name):
        super().__init__()
        self.file1 = file1
        self.file2 = file2
        self.sheet1_name = sheet1_name
        self.sheet2_name = sheet2_name

    def run(self):
        try:
            # Load with styles
            df1, raw_styles1, merges1, formulas1, col_widths1, row_heights1 = load_excel_with_styles(self.file1, self.sheet1_name)
            df2, raw_styles2, merges2, formulas2, col_widths2, row_heights2 = load_excel_with_styles(self.file2, self.sheet2_name)
            
            result = compare_dataframes(df1, df2)
            
            self.finished.emit(df1, df2, raw_styles1, raw_styles2, merges1, merges2, formulas1, formulas2, col_widths1, row_heights1, col_widths2, row_heights2, result)
        except Exception as e:
            self.error.emit(str(e) + "\n" + traceback.format_exc())

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Document Compare AI") # Rebranded
        self.resize(1280, 850)
        self.setMinimumSize(800, 600)
        
        # Central Widget
        self.central_widget = QWidget()
        self.central_widget.setObjectName("AppRoot")
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)
        if getattr(sys, 'frozen', False):
            # Running as compiled executable
            base_path = Path(sys._MEIPASS)
            self.icon_dir = base_path / "ui" / "assets" / "icons"
            logo_assets = base_path / "ui" / "assets"
            logo_root = base_path / "ui"
        else:
            # Running source
            ui_root = Path(__file__).resolve().parent
            self.icon_dir = ui_root / "assets" / "icons"
            if not self.icon_dir.exists():
                self.icon_dir = ui_root / ".." / "assets" / "icons"
            logo_assets = ui_root / "assets"
            if not logo_assets.exists():
                logo_assets = ui_root / ".." / "assets"
            logo_root = ui_root

        self.logo_paths = [
            logo_assets / "logo.png",
            logo_assets / "logo.jpg",
            logo_root / "logo.png",
            logo_root / "logo.jpg",
        ]

        # 0. Ribbon
        self.init_ribbon()

        # 1. Header (Mode + Branding)
        self.init_header()
        
        # State
        self.file1_path = None
        self.file2_path = None
        self.sheet1_name = None
        self.sheet2_name = None
        self.sheet1_names = []
        self.sheet2_names = []
        self.last_open_dir = get_last_open_dir()
        self.file1_converted_path = None
        self.file2_converted_path = None

        # 2. Content Area (Stacked Pages)
        self.stack = QStackedWidget()
        self.main_layout.addWidget(self.stack, 1)
        self.init_footer()

        # Init Pages
        self.init_excel_page()
        self.init_word_page() # Placeholder
        self.init_ppt_page()  # Placeholder
        
        self.result_view = ResultView(self.go_back)
        self.result_view.setObjectName("ResultView")
        self.result_view.swapRequested.connect(self.handle_result_swap)
        self.result_view.sheetChangeRequested.connect(self.on_result_sheet_change)
        self.stack.addWidget(self.result_view)

        self.connect_ribbon_actions()
        self.stack.currentChanged.connect(self.on_stack_changed)
        self.on_stack_changed(self.stack.currentIndex())
        
        # Default to Excel
        self.mode_group.button(0).setChecked(True)
        self.apply_global_styles()

        self.update_info = None
        self.update_check_thread = None

        self.handle_post_update()
        QTimer.singleShot(200, self.startup_update_check)

    def init_ribbon(self):
        self.ribbon = QFrame()
        self.ribbon.setObjectName("Ribbon")
        ribbon_layout = QVBoxLayout(self.ribbon)
        ribbon_layout.setContentsMargins(12, 2, 12, 2)
        ribbon_layout.setSpacing(0)

        self.ribbon_tabs = QTabWidget()
        self.ribbon_tabs.setObjectName("RibbonTabs")
        self.ribbon_tabs.setUsesScrollButtons(False)
        self.ribbon_tabs.setTabPosition(QTabWidget.TabPosition.North)

        home_tab = QWidget()
        self.ribbon_home_tab = home_tab
        home_layout = QVBoxLayout(home_tab)
        home_layout.setContentsMargins(6, 2, 6, 2)
        home_layout.setSpacing(0)

        self.ribbon_home_panel = QWidget()
        self.ribbon_home_panel.setObjectName("RibbonPanel")
        home_panel_layout = QHBoxLayout(self.ribbon_home_panel)
        home_panel_layout.setContentsMargins(0, 0, 0, 0)
        home_panel_layout.setSpacing(8)
        home_panel_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        self.ribbon_home_scroll = QScrollArea()
        self.ribbon_home_scroll.setObjectName("RibbonScroll")
        self.ribbon_home_scroll.setWidgetResizable(True)
        self.ribbon_home_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.ribbon_home_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.ribbon_home_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.ribbon_home_scroll.setWidget(self.ribbon_home_panel)
        home_layout.addWidget(self.ribbon_home_scroll)

        icon_size = QSize(18, 18)

        # Navigation Buttons (Back)
        nav_layout = QHBoxLayout()
        nav_layout.setContentsMargins(0, 0, 0, 0)
        nav_layout.setSpacing(6)
        nav_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        self.ribbon_back_btn = QToolButton()
        self.ribbon_back_btn.setObjectName("RibbonButton")
        self.ribbon_back_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_back_btn.setText("Back")
        self.ribbon_back_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowBack))
        self.ribbon_back_btn.setIconSize(icon_size)
        nav_layout.addWidget(self.ribbon_back_btn)
        home_panel_layout.addLayout(nav_layout)

        # Separator
        sep_nav = QFrame()
        sep_nav.setObjectName("RibbonSeparator")
        sep_nav.setFrameShape(QFrame.Shape.VLine)
        sep_nav.setFrameShadow(QFrame.Shadow.Sunken)
        home_panel_layout.addWidget(sep_nav)

        # Number Buttons
        number_layout = QHBoxLayout()
        number_layout.setContentsMargins(0, 0, 0, 0)
        number_layout.setSpacing(6)
        number_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        self.ribbon_dec_up_btn = QToolButton()
        self.ribbon_dec_up_btn.setObjectName("RibbonButton")
        self.ribbon_dec_up_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_dec_up_btn.setText("Increase")
        self.ribbon_dec_up_btn.setIcon(self.load_icon("IncreaseDecimal-03295B", QStyle.StandardPixmap.SP_ArrowUp))
        self.ribbon_dec_up_btn.setIconSize(icon_size)
        number_layout.addWidget(self.ribbon_dec_up_btn)

        self.ribbon_dec_down_btn = QToolButton()
        self.ribbon_dec_down_btn.setObjectName("RibbonButton")
        self.ribbon_dec_down_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_dec_down_btn.setText("Decrease")
        self.ribbon_dec_down_btn.setIcon(self.load_icon("DecreaseDecimal-03295B", QStyle.StandardPixmap.SP_ArrowDown))
        self.ribbon_dec_down_btn.setIconSize(icon_size)
        number_layout.addWidget(self.ribbon_dec_down_btn)
        home_panel_layout.addLayout(number_layout)

        sep1 = QFrame()
        sep1.setObjectName("RibbonSeparator")
        sep1.setFrameShape(QFrame.Shape.VLine)
        sep1.setFrameShadow(QFrame.Shadow.Sunken)
        home_panel_layout.addWidget(sep1)

        # Cells Buttons
        cells_layout = QHBoxLayout()
        cells_layout.setContentsMargins(0, 0, 0, 0)
        cells_layout.setSpacing(6)
        cells_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        self.ribbon_auto_fit_btn = QToolButton()
        self.ribbon_auto_fit_btn.setObjectName("RibbonButton")
        self.ribbon_auto_fit_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_auto_fit_btn.setText("Auto-Fit")
        self.ribbon_auto_fit_btn.setIcon(self.load_icon("auto-fit", QStyle.StandardPixmap.SP_FileDialogDetailedView))
        self.ribbon_auto_fit_btn.setIconSize(icon_size)
        cells_layout.addWidget(self.ribbon_auto_fit_btn)

        home_panel_layout.addLayout(cells_layout)

        sep2 = QFrame()
        sep2.setObjectName("RibbonSeparator")
        sep2.setFrameShape(QFrame.Shape.VLine)
        sep2.setFrameShadow(QFrame.Shadow.Sunken)
        home_panel_layout.addWidget(sep2)

        # Compare Buttons
        compare_layout = QHBoxLayout()
        compare_layout.setContentsMargins(0, 0, 0, 0)
        compare_layout.setSpacing(6)
        compare_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        self.ribbon_diff_btn = QToolButton()
        self.ribbon_diff_btn.setObjectName("RibbonButton")
        self.ribbon_diff_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_diff_btn.setText("Differences")
        self.ribbon_diff_btn.setCheckable(True)
        self.ribbon_diff_btn.setChecked(False)
        self.ribbon_diff_btn.setIcon(self.load_icon("diff-toggle", QStyle.StandardPixmap.SP_DialogApplyButton))
        self.ribbon_diff_btn.setIconSize(icon_size)
        compare_layout.addWidget(self.ribbon_diff_btn)

        self.ribbon_diff_content_btn = QToolButton()
        self.ribbon_diff_content_btn.setObjectName("RibbonButton")
        self.ribbon_diff_content_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_diff_content_btn.setText("Content")
        self.ribbon_diff_content_btn.setCheckable(True)
        self.ribbon_diff_content_btn.setChecked(True)
        self.ribbon_diff_content_btn.setIcon(self.load_icon("diff-content", QStyle.StandardPixmap.SP_FileDialogContentsView))
        self.ribbon_diff_content_btn.setIconSize(icon_size)
        compare_layout.addWidget(self.ribbon_diff_content_btn)

        self.ribbon_diff_format_btn = QToolButton()
        self.ribbon_diff_format_btn.setObjectName("RibbonButton")
        self.ribbon_diff_format_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_diff_format_btn.setText("Format")
        self.ribbon_diff_format_btn.setCheckable(True)
        self.ribbon_diff_format_btn.setChecked(True)
        self.ribbon_diff_format_btn.setIcon(self.load_icon("diff-format", QStyle.StandardPixmap.SP_FileDialogInfoView))
        self.ribbon_diff_format_btn.setIconSize(icon_size)
        compare_layout.addWidget(self.ribbon_diff_format_btn)

        self.ribbon_diff_formula_btn = QToolButton()
        self.ribbon_diff_formula_btn.setObjectName("RibbonButton")
        self.ribbon_diff_formula_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_diff_formula_btn.setText("Formula")
        self.ribbon_diff_formula_btn.setCheckable(True)
        self.ribbon_diff_formula_btn.setChecked(False)
        self.ribbon_diff_formula_btn.setIcon(self.load_icon("diff-formula", QStyle.StandardPixmap.SP_FileDialogInfoView))
        self.ribbon_diff_formula_btn.setIconSize(icon_size)
        compare_layout.addWidget(self.ribbon_diff_formula_btn)

        self.ribbon_settings_btn = QToolButton()
        self.ribbon_settings_btn.setObjectName("RibbonButton")
        self.ribbon_settings_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_settings_btn.setText("Settings")
        self.ribbon_settings_btn.setIcon(self.load_icon("settings", QStyle.StandardPixmap.SP_FileDialogDetailedView))
        self.ribbon_settings_btn.setIconSize(icon_size)
        self.ribbon_settings_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)

        settings_menu = QMenu(self)
        self.ribbon_ignore_case_action = QAction("Ignore Case", self)
        self.ribbon_ignore_case_action.setCheckable(True)
        self.ribbon_ignore_case_action.setChecked(False)
        settings_menu.addAction(self.ribbon_ignore_case_action)

        self.ribbon_ignore_space_action = QAction("Ignore Space", self)
        self.ribbon_ignore_space_action.setCheckable(True)
        self.ribbon_ignore_space_action.setChecked(False)
        settings_menu.addAction(self.ribbon_ignore_space_action)

        self.ribbon_ignore_numfmt_action = QAction("Ignore NumFmt", self)
        self.ribbon_ignore_numfmt_action.setCheckable(True)
        self.ribbon_ignore_numfmt_action.setChecked(False)
        settings_menu.addAction(self.ribbon_ignore_numfmt_action)

        self.ribbon_settings_btn.setMenu(settings_menu)
        compare_layout.addWidget(self.ribbon_settings_btn)

        self.ribbon_only_changes_btn = QToolButton()
        self.ribbon_only_changes_btn.setObjectName("RibbonButton")
        self.ribbon_only_changes_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_only_changes_btn.setText("OnlyChanges")
        self.ribbon_only_changes_btn.setCheckable(True)
        self.ribbon_only_changes_btn.setChecked(False)
        self.ribbon_only_changes_btn.setIcon(self.load_icon("diff-filter", QStyle.StandardPixmap.SP_DirIcon))
        self.ribbon_only_changes_btn.setIconSize(icon_size)
        compare_layout.addWidget(self.ribbon_only_changes_btn)
        home_panel_layout.addLayout(compare_layout)
        # Ensure the scroll area has a stable content size so buttons stay visible.
        self.ribbon_home_panel.setMinimumWidth(self.ribbon_home_panel.sizeHint().width())
        self.ribbon_home_panel.setMinimumHeight(self.ribbon_home_panel.sizeHint().height())

        # View Tab
        view_tab = QWidget()
        self.ribbon_view_tab = view_tab
        view_layout = QHBoxLayout(view_tab)
        view_layout.setContentsMargins(8, 2, 8, 2)
        view_layout.setSpacing(8)
        view_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        # View Mode Buttons
        view_buttons_layout = QHBoxLayout()
        view_buttons_layout.setContentsMargins(0, 0, 0, 0)
        view_buttons_layout.setSpacing(6)
        view_buttons_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        self.ribbon_view_group = QButtonGroup(self)
        self.ribbon_view_group.setExclusive(True)

        self.ribbon_view_normal_btn = QToolButton()
        self.ribbon_view_normal_btn.setObjectName("RibbonButton")
        self.ribbon_view_normal_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_view_normal_btn.setText("Normal")
        self.ribbon_view_normal_btn.setCheckable(True)
        self.ribbon_view_normal_btn.setChecked(True)
        # Load state-aware icon
        icon_normal = QIcon()
        icon_opt_off = self.load_svg_colored_icon("view-normal", "#64748b", size=18)
        icon_opt_on = self.load_svg_colored_icon("view-normal", "#ffffff", size=18)

        def _icon_pixmap(icon_obj, size):
            if icon_obj is None:
                return None
            if isinstance(icon_obj, QPixmap):
                return icon_obj
            if isinstance(icon_obj, QIcon):
                return icon_obj.pixmap(size, size)
            return None
        
        pix_normal_off = _icon_pixmap(icon_opt_off, 18)
        pix_normal_on = _icon_pixmap(icon_opt_on, 18)
        if pix_normal_off:
            icon_normal.addPixmap(pix_normal_off, QIcon.Mode.Normal, QIcon.State.Off)
        if pix_normal_on:
            icon_normal.addPixmap(pix_normal_on, QIcon.Mode.Normal, QIcon.State.On)
            icon_normal.addPixmap(pix_normal_on, QIcon.Mode.Active, QIcon.State.On)

        self.ribbon_view_normal_btn.setIcon(icon_normal)
        self.ribbon_view_normal_btn.setIconSize(icon_size)
        self.ribbon_view_group.addButton(self.ribbon_view_normal_btn, 0)
        view_buttons_layout.addWidget(self.ribbon_view_normal_btn)

        self.ribbon_view_page_btn = QToolButton()
        self.ribbon_view_page_btn.setObjectName("RibbonButton")
        self.ribbon_view_page_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_view_page_btn.setText("Page Break")
        self.ribbon_view_page_btn.setCheckable(True)
        # Load state-aware icon
        icon_page = QIcon()
        icon_page_off = self.load_svg_colored_icon("view-page-break", "#64748b", size=18)
        icon_page_on = self.load_svg_colored_icon("view-page-break", "#ffffff", size=18)

        pix_page_off = _icon_pixmap(icon_page_off, 18)
        pix_page_on = _icon_pixmap(icon_page_on, 18)
        if pix_page_off:
            icon_page.addPixmap(pix_page_off, QIcon.Mode.Normal, QIcon.State.Off)
        if pix_page_on:
            icon_page.addPixmap(pix_page_on, QIcon.Mode.Normal, QIcon.State.On)
            icon_page.addPixmap(pix_page_on, QIcon.Mode.Active, QIcon.State.On)
             
        self.ribbon_view_page_btn.setIcon(icon_page)
        self.ribbon_view_page_btn.setIconSize(icon_size)
        self.ribbon_view_group.addButton(self.ribbon_view_page_btn, 1)
        view_buttons_layout.addWidget(self.ribbon_view_page_btn)
        
        view_layout.addLayout(view_buttons_layout)

        view_layout.addStretch()

        # Help Tab
        help_tab = QWidget()
        self.ribbon_help_tab = help_tab
        help_layout = QHBoxLayout(help_tab)
        help_layout.setContentsMargins(8, 6, 8, 6)
        help_layout.setSpacing(8)
        help_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        help_buttons = QHBoxLayout()
        help_buttons.setContentsMargins(0, 0, 0, 0)
        help_buttons.setSpacing(6)
        help_buttons.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        self.ribbon_feedback_btn = QToolButton()
        self.ribbon_feedback_btn.setObjectName("RibbonButton")
        self.ribbon_feedback_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_feedback_btn.setText("Feedback")
        self.ribbon_feedback_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MessageBoxInformation))
        self.ribbon_feedback_btn.setIconSize(icon_size)
        help_buttons.addWidget(self.ribbon_feedback_btn)

        self.ribbon_about_btn = QToolButton()
        self.ribbon_about_btn.setObjectName("RibbonButton")
        self.ribbon_about_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_about_btn.setText("About")
        self.ribbon_about_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MessageBoxQuestion))
        self.ribbon_about_btn.setIconSize(icon_size)
        self.ribbon_about_btn.clicked.connect(self.show_about_dialog)
        help_buttons.addWidget(self.ribbon_about_btn)

        self.ribbon_update_btn = QToolButton()
        self.ribbon_update_btn.setObjectName("RibbonButton")
        self.ribbon_update_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.ribbon_update_btn.setText("Update")
        self.ribbon_update_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_BrowserReload))
        self.ribbon_update_btn.setIconSize(icon_size)
        self.ribbon_update_btn.clicked.connect(self.check_for_updates)
        help_buttons.addWidget(self.ribbon_update_btn)

        help_layout.addLayout(help_buttons)
        help_layout.addStretch()

        # Add Tabs (Removed File and Formula)
        self.ribbon_tabs.addTab(home_tab, "Home")
        self.ribbon_tabs.addTab(view_tab, "View")
        self.ribbon_tabs.addTab(help_tab, "Help")
        self.ribbon_tabs.setCurrentIndex(0) # Adjusted index after removing tabs

        tabs_row = QHBoxLayout()
        tabs_row.setContentsMargins(0, 0, 0, 0)
        tabs_row.setSpacing(6)
        tabs_row.addWidget(self.ribbon_tabs, 1)

        self.ribbon_pin_btn = QToolButton()
        self.ribbon_pin_btn.setObjectName("RibbonPin")
        self.ribbon_pin_btn.setCheckable(False)
        self.ribbon_pin_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_TitleBarShadeButton))
        self.ribbon_pin_btn.setToolTip("Ribbon: Groups")
        self.ribbon_pin_btn.clicked.connect(self.cycle_ribbon_mode)
        self.ribbon_mode = "groups"  # groups, hidden
        tabs_row.addWidget(self.ribbon_pin_btn, 0, Qt.AlignmentFlag.AlignRight)

        ribbon_layout.addLayout(tabs_row)
        self.main_layout.addWidget(self.ribbon)

        self._toggle_icons = {}
        self.setup_toggle_icons()
        self.update_ribbon_sizes()

    def init_header(self):
        self.header = QFrame()
        self.header.setObjectName("TopBar")
        header_layout = QHBoxLayout(self.header)
        header_layout.setContentsMargins(24, 8, 24, 8)
        header_layout.setSpacing(12)

        brand_container = QWidget()
        brand_layout = QVBoxLayout(brand_container)
        brand_layout.setContentsMargins(0, 0, 0, 0)
        brand_layout.setSpacing(2)
        title_label = QLabel("DocCompare AI")
        title_label.setObjectName("AppTitle")
        subtitle_label = QLabel("Smart document diff, focused on clarity")
        subtitle_label.setObjectName("AppTagline")
        brand_layout.addWidget(title_label)
        brand_layout.addWidget(subtitle_label)
        header_layout.addWidget(brand_container)

        header_layout.addStretch()

        mode_container = QFrame()
        mode_container.setObjectName("ModePill")
        mode_layout = QHBoxLayout(mode_container)
        mode_layout.setContentsMargins(4, 4, 4, 4)
        mode_layout.setSpacing(4)

        self.mode_group = QButtonGroup(self)
        self.mode_group.setExclusive(True)
        self.mode_buttons = []

        modes = ["Excel", "Word (Beta)", "PowerPoint (Beta)"]
        for idx, label in enumerate(modes):
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setProperty("modeTab", True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            self.mode_group.addButton(btn, idx)
            mode_layout.addWidget(btn)
            self.mode_buttons.append(btn)

        self.mode_group.idClicked.connect(self.change_page)
        header_layout.addWidget(mode_container)

        header_layout.addStretch()

        self.logo_label = QLabel()
        self.logo_label.setObjectName("LogoLabel")
        self.logo_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.apply_logo()
        header_layout.addWidget(self.logo_label)

        self.main_layout.addWidget(self.header)

    def init_footer(self):
        self.footer = QFrame()
        self.footer.setObjectName("FooterBar")
        footer_layout = QHBoxLayout(self.footer)
        footer_layout.setContentsMargins(16, 4, 16, 8)
        footer_layout.setSpacing(8)
        footer_layout.addStretch()

        self.version_label = ClickableLabel(f"v{APP_VERSION} Desktop")
        self.version_label.setObjectName("VersionLabel")
        self.version_label.clicked.connect(self.on_version_label_clicked)
        footer_layout.addWidget(self.version_label)

        self.main_layout.addWidget(self.footer)

    def change_page(self, index):
        # 0 -> Excel, 1 -> Word, 2 -> PPT. Result View is index 3 (manual jump).
        if index < 3:
             self.stack.setCurrentIndex(index)

    def init_excel_page(self):
        page = QWidget()
        page.setObjectName("ExcelPage")
        layout = QVBoxLayout(page)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(24)
        
        # Excel Styling
        title = QLabel("Excel Comparison")
        title.setObjectName("PageTitle")
        layout.addWidget(title)

        subtitle = QLabel("Compare values, formulas, and formatting in a clean side-by-side view.")
        subtitle.setObjectName("PageSubtitle")
        layout.addWidget(subtitle)

        chips = QHBoxLayout()
        chips.setSpacing(8)
        for text in ["Values", "Formulas", "Formatting"]:
            chip = QLabel(text)
            chip.setProperty("chip", True)
            chips.addWidget(chip)
        chips.addStretch()
        layout.addLayout(chips)

        card = QFrame()
        card.setObjectName("UploadCard")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(24, 24, 24, 24)
        card_layout.setSpacing(16)

        card_title = QLabel("Start a new comparison")
        card_title.setObjectName("CardTitle")
        card_layout.addWidget(card_title)

        card_subtitle = QLabel("Drag two Excel files here, or click a card to browse.")
        card_subtitle.setObjectName("CardSubtitle")
        card_layout.addWidget(card_subtitle)

        drop_layout = QHBoxLayout()
        drop_layout.setSpacing(16)

        self.drop1 = FileDropWidget("Original Excel", "#14b8a6")
        self.drop1.fileDropped.connect(self.set_file1)
        self.drop1.filesDropped.connect(self.handle_files_drop)
        self.drop1.set_initial_dir(self.last_open_dir)
        
        self.drop2 = FileDropWidget("Modified Excel", "#f97316")
        self.drop2.fileDropped.connect(self.set_file2)
        self.drop2.filesDropped.connect(self.handle_files_drop)
        self.drop2.set_initial_dir(self.last_open_dir)

        drop_layout.addWidget(self.drop1, 1)

        # Swap Button
        self.swap_btn = QToolButton()
        self.swap_btn.setObjectName("SwapButton")
        self.swap_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_BrowserReload)) 
        self.swap_btn.setToolTip("Swap Files")
        self.swap_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.swap_btn.clicked.connect(self.swap_inputs)
        self.swap_btn.setStyleSheet("""
            QToolButton#SwapButton {
                background-color: #f1f5f9;
                border: 1px solid #e2e8f0;
                border-radius: 16px;
                padding: 6px;
                min-width: 32px;
                min-height: 32px;
            }
            QToolButton#SwapButton:hover {
                background-color: #e2e8f0;
                border-color: #cbd5e1;
            }
            QToolButton#SwapButton:pressed {
                background-color: #cbd5e1;
            }
        """)
        drop_layout.addWidget(self.swap_btn, 0, Qt.AlignmentFlag.AlignCenter)

        drop_layout.addWidget(self.drop2, 1)
        card_layout.addLayout(drop_layout)

        sheet_layout = QHBoxLayout()
        sheet_layout.setSpacing(16)

        self.sheet1_combo = QComboBox()
        self.sheet1_combo.setObjectName("SheetCombo")
        self.sheet1_combo.setEnabled(False)
        self.sheet1_combo.currentTextChanged.connect(self.on_sheet1_changed)

        sheet1_row = QHBoxLayout()
        sheet1_label = QLabel("Sheet")
        sheet1_label.setObjectName("SheetLabel")
        sheet1_row.addWidget(sheet1_label)
        sheet1_row.addWidget(self.sheet1_combo, 1)
        sheet1_container = QWidget()
        sheet1_container.setLayout(sheet1_row)

        self.sheet2_combo = QComboBox()
        self.sheet2_combo.setObjectName("SheetCombo")
        self.sheet2_combo.setEnabled(False)
        self.sheet2_combo.currentTextChanged.connect(self.on_sheet2_changed)

        sheet2_row = QHBoxLayout()
        sheet2_label = QLabel("Sheet")
        sheet2_label.setObjectName("SheetLabel")
        sheet2_row.addWidget(sheet2_label)
        sheet2_row.addWidget(self.sheet2_combo, 1)
        sheet2_container = QWidget()
        sheet2_container.setLayout(sheet2_row)

        sheet_layout.addWidget(sheet1_container, 1)
        sheet_layout.addWidget(sheet2_container, 1)
        card_layout.addLayout(sheet_layout)

        action_layout = QHBoxLayout()
        action_layout.setSpacing(12)
        self.helper_label = QLabel("Supports .xlsx and .xls files")
        self.helper_label.setObjectName("HelperText")
        action_layout.addWidget(self.helper_label)
        action_layout.addStretch()

        self.compare_btn = QPushButton("Compare Now")
        self.compare_btn.setObjectName("PrimaryButton")
        self.compare_btn.clicked.connect(self.start_comparison)
        self.compare_btn.setEnabled(False)
        action_layout.addWidget(self.compare_btn)

        card_layout.addLayout(action_layout)
        layout.addWidget(card)
        layout.addStretch()
        
        self.stack.addWidget(page)
        
    def init_word_page(self):
        page = self.create_coming_soon_page("Word Document Comparison", "Coming Soon")
        self.stack.addWidget(page)

    def init_ppt_page(self):
        page = self.create_coming_soon_page("PowerPoint Presentation Comparison", "Coming Soon")
        self.stack.addWidget(page)
        
    def create_coming_soon_page(self, title_text, status_text):
        page = QWidget()
        page.setObjectName("ComingSoonPage")
        layout = QVBoxLayout(page)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        title = QLabel(title_text)
        title.setObjectName("ComingSoonTitle")
        layout.addWidget(title)
        
        status = QLabel(status_text)
        status.setObjectName("ComingSoonStatus")
        layout.addWidget(status)
        
        return page

    def handle_files_drop(self, files):
        if len(files) >= 2:
            self.set_file1(files[0])
            self.set_file2(files[1])

    def swap_inputs(self):
        self.swap_btn.setEnabled(False)
        
        # Store current state
        f1, f2 = self.file1_path, self.file2_path
        s1, s2 = self.sheet1_name, self.sheet2_name
        
        # Swap logic
        # Note: process_file emits signals that call set_file1/2
        
        # Swap File 1 (takes old File 2)
        if f2:
            self.drop1.process_file(f2)
        else:
            self.drop1.clear_file()
            
        # Swap File 2 (takes old File 1)
        if f1:
            self.drop2.process_file(f1)
        else:
            self.drop2.clear_file()
            
        # Attempt to restore sheet selections
        # We want the new Left items (File 2) to use the old Right selection (s2)
        if s2 and self.sheet1_combo.isEnabled():
            self.sheet1_combo.setCurrentText(s2)
        
        # We want the new Right items (File 1) to use the old Left selection (s1)
        if s1 and self.sheet2_combo.isEnabled():
            self.sheet2_combo.setCurrentText(s1)
            
        self.swap_btn.setEnabled(True)

    def handle_result_swap(self):
        # Swap inputs
        self.swap_inputs()
        # Trigger comparison again if ready
        if self.compare_btn.isEnabled():
            self.start_comparison()


    def set_file1(self, path):
        self.file1_path = path # Original path for UI
        self.file1_converted_path = None
        
        path_lower = path.lower()
        if path_lower.endswith(".xls") and not path_lower.endswith(".xlsx"):
            # Attempt convert
            converted = convert_xls_to_xlsx(path)
            if converted:
                self.file1_converted_path = converted
                # Use converted path for processing/loading sheets
                # But keep original path for display in UI
            else:
                 QMessageBox.warning(self, "Format Warning", 
                                   "This is a legacy .xls file and requires 'xlrd' to process.\n"
                                   "Values will be compared, but some styles/formulas might be lost during conversion.")
        
        self.update_last_dir(path)
        # Use converted path for internal processing if available
        proc_path = self.file1_converted_path if self.file1_converted_path else path
        self.populate_sheet_combo(self.sheet1_combo, proc_path, is_left=True)
        self.check_ready()

    def set_file2(self, path):
        self.file2_path = path # Original path for UI
        self.file2_converted_path = None
        
        path_lower = path.lower()
        if path_lower.endswith(".xls") and not path_lower.endswith(".xlsx"):
            converted = convert_xls_to_xlsx(path)
            if converted:
                self.file2_converted_path = converted
            else:
                 QMessageBox.warning(self, "Format Warning", 
                                   "This is a legacy .xls file and requires 'xlrd' to process.\n"
                                   "Values will be compared, but some styles/formulas might be lost during conversion.")

        self.update_last_dir(path)
        proc_path = self.file2_converted_path if self.file2_converted_path else path
        self.populate_sheet_combo(self.sheet2_combo, proc_path, is_left=False)
        self.check_ready()
        
    def check_ready(self):
        if self.file1_path and self.file2_path and self.sheet1_name and self.sheet2_name:
            self.compare_btn.setEnabled(True)
            self.compare_btn.setText("Compare Now")
            self.helper_label.setText("Ready to analyze.")
        else:
            self.compare_btn.setEnabled(False)
            self.compare_btn.setText("Compare Now")
            self.helper_label.setText("Select two files to compare.")

    def start_comparison(self):
        # ... (UI updates) ...
        f1 = self.file1_converted_path if self.file1_converted_path else self.file1_path
        f2 = self.file2_converted_path if self.file2_converted_path else self.file2_path
        
        self.result_view.set_loading(True)
        
        self.worker = WorkerThread(f1, f2, self.sheet1_name, self.sheet2_name)
        self.worker.finished.connect(self.on_comparison_finished)
        self.worker.error.connect(self.on_error)
        self.worker.start()
        
    def on_comparison_finished(self, df1, df2, s1, s2, m1, m2, f1, f2, cw1, rh1, cw2, rh2, result):
        # When finished, we jump to Result View
        # Result View works with data, so we populate it
        self.result_view.load_data(df1, df2, s1, s2, m1, m2, f1, f2, cw1, rh1, cw2, rh2, result, self.file1_path, self.file2_path)
        self.result_view.set_loading(False)
        self.result_view.set_sheet_options(1, self.sheet1_names or [], self.sheet1_name)
        self.result_view.set_sheet_options(2, self.sheet2_names or [], self.sheet2_name)
        self.stack.setCurrentWidget(self.result_view)
        
        # Keep Excel mode selected when showing results
        self.mode_group.button(0).setChecked(True)
        
        self.compare_btn.setText("Compare Now")
        self.compare_btn.setEnabled(True)
    
    def on_error(self, err_msg):
        QMessageBox.critical(self, "Error", err_msg)
        self.compare_btn.setText("Compare Now")
        self.compare_btn.setEnabled(True)
        self.result_view.set_loading(False)

    def go_back(self):
        self.stack.setCurrentIndex(0) # Back to Excel upload

    def update_diff_controls_state(self, enabled):
        self.ribbon_diff_content_btn.setEnabled(enabled)
        self.ribbon_diff_format_btn.setEnabled(enabled)
        self.ribbon_diff_formula_btn.setEnabled(enabled)
        self.ribbon_only_changes_btn.setEnabled(enabled)
        self.ribbon_settings_btn.setEnabled(enabled)
        
        # Visual cues for opacity if needed, though SetEnabled handles most
        opacity = 1.0 if enabled else 0.5
        for btn in [self.ribbon_diff_content_btn, self.ribbon_diff_format_btn, 
                   self.ribbon_diff_formula_btn, self.ribbon_only_changes_btn, self.ribbon_settings_btn]:
            effect = btn.graphicsEffect()
            if not effect:
                from PyQt6.QtWidgets import QGraphicsOpacityEffect
                effect = QGraphicsOpacityEffect(btn)
                btn.setGraphicsEffect(effect)
            effect.setOpacity(opacity)

    def connect_ribbon_actions(self):
        self.ribbon_back_btn.clicked.connect(self.go_back)
        self.ribbon_dec_down_btn.clicked.connect(lambda: self.result_view.adjust_decimals(-1))
        self.ribbon_dec_up_btn.clicked.connect(lambda: self.result_view.adjust_decimals(1))
        self.ribbon_auto_fit_btn.clicked.connect(self.result_view.auto_fit_columns)
        self.ribbon_diff_btn.toggled.connect(self.result_view.set_show_differences)
        self.ribbon_diff_btn.toggled.connect(self.update_diff_controls_state)
        # Initialize state
        self.update_diff_controls_state(self.ribbon_diff_btn.isChecked())
        self.ribbon_diff_content_btn.toggled.connect(lambda v: self.result_view.set_diff_options(content=v))
        self.ribbon_diff_format_btn.toggled.connect(lambda v: self.result_view.set_diff_options(format=v))
        self.ribbon_only_changes_btn.toggled.connect(lambda v: self.result_view.set_diff_options(only_changes=v))
        self.ribbon_diff_formula_btn.toggled.connect(lambda v: self.result_view.set_diff_options(formula=v))
        self.ribbon_ignore_case_action.toggled.connect(lambda v: self.result_view.set_diff_options(ignore_case=v))
        self.ribbon_ignore_space_action.toggled.connect(lambda v: self.result_view.set_diff_options(ignore_whitespace=v))
        self.ribbon_ignore_numfmt_action.toggled.connect(lambda v: self.result_view.set_diff_options(ignore_number_format=v))
        if hasattr(self, "ribbon_view_group"):
            self.ribbon_view_group.idClicked.connect(
                lambda idx: self.result_view.set_view_mode("normal" if idx == 0 else "page_break")
            )
            self.result_view.viewModeChanged.connect(self.on_result_view_mode_changed)

        if hasattr(self, "ribbon_feedback_btn"):
            self.ribbon_feedback_btn.clicked.connect(self.open_feedback_dialog)

    def cycle_ribbon_mode(self):
        order = ["groups", "hidden"]
        try:
            idx = order.index(self.ribbon_mode)
        except ValueError:
            idx = 0
        self.ribbon_mode = order[(idx + 1) % len(order)]
        self.apply_ribbon_mode()

    def load_icon(self, name, fallback):
        # Try both svg and png
        for ext in [".svg", ".png"]:
            path = self.icon_dir / f"{name}{ext}"
            if path.exists():
                return QIcon(str(path))
        return self.style().standardIcon(fallback)

    def open_feedback_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Feedback")
        dialog.setMinimumWidth(420)

        layout = QVBoxLayout(dialog)
        title = QLabel("We value your feedback")
        title.setObjectName("ResultTitle")
        layout.addWidget(title)

        note = QLabel("Tell us what you liked or what we should improve.")
        note.setStyleSheet("color: #64748b; font-size: 12px;")
        layout.addWidget(note)

        input_box = QTextEdit()
        input_box.setPlaceholderText("Write your feedback here...")
        layout.addWidget(input_box)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Cancel | QDialogButtonBox.StandardButton.Ok)
        buttons.button(QDialogButtonBox.StandardButton.Ok).setText("Submit")
        buttons.accepted.connect(lambda: self.save_feedback(dialog, input_box.toPlainText()))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        dialog.exec()

    def show_about_dialog(self):
        self.show_author_info()

    def show_author_info(self):
        title = "About DocCompare AI"
        text = (
            f"<b>DocCompare AI v{APP_VERSION}</b><br><br>"
            "Created by: <b>KEI AI Solutions</b><br>"
            "Contact: <b>Khoa.Le</b><br><br>"
            "This tool is free to use. No license required.<br>"
            "Sharing and feedback are welcome!"
        )
        QMessageBox.about(self, title, text)

    def apply_logo(self):
        if hasattr(self, "logo_label") and getattr(self, "logo_paths", None):
            for logo_path in self.logo_paths:
                if not logo_path.exists():
                    continue
                pix = QPixmap(str(logo_path))
                if pix.isNull():
                    continue
                max_h = 90  # Target height
                max_w = 600
                scaled = pix.scaled(
                    max_w,
                    max_h,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                )
                
                # Composite onto white background to ensure visibility and clarity
                # This fixes "edges lost" if it was due to transparency/rendering artifacts
                # and satisfies "white background" request.
                final_pix = QPixmap(scaled.size())
                final_pix.fill(Qt.GlobalColor.white)
                painter = QPainter(final_pix)
                painter.drawPixmap(0, 0, scaled)
                painter.end()

                self.logo_label.setText("")
                self.logo_label.setPixmap(final_pix)
                self.logo_label.setProperty("placeholder", False)
                self.logo_label.setToolTip("")
                self.logo_label.style().unpolish(self.logo_label)
                self.logo_label.style().polish(self.logo_label)
                return
        if hasattr(self, "logo_label"):
            self.logo_label.setPixmap(QPixmap())
            self.logo_label.setText("LOGO")
            self.logo_label.setProperty("placeholder", True)
            self.logo_label.setToolTip("Place logo at ui/assets/logo.png (or ui/logo.png)")
            self.logo_label.style().unpolish(self.logo_label)
            self.logo_label.style().polish(self.logo_label)

    def check_for_updates(self):
        self.start_update_check(show_prompt=True, manual=True)

    def startup_update_check(self):
        self.start_update_check(show_prompt=get_update_prompt_on_startup(), manual=False)

    def start_update_check(self, show_prompt=False, manual=False):
        if self.update_check_thread and self.update_check_thread.isRunning():
            return
        if manual:
            self.setCursor(Qt.CursorShape.WaitCursor)
        self.update_check_thread = UpdateCheckThread(CURRENT_VERSION)
        self.update_check_thread.finished.connect(
            lambda info, err, sp=show_prompt, man=manual: self.handle_update_result(info, err, sp, man)
        )
        self.update_check_thread.start()

    def handle_update_result(self, info, err, show_prompt, manual):
        if manual:
            self.setCursor(Qt.CursorShape.ArrowCursor)

        if err:
            self.update_info = None
            self.update_version_label(status="unknown")
            if manual:
                QMessageBox.warning(self, "Check for Updates", f"Unable to check updates.\n\n{err}")
            else:
                 # Startup check failed -> possibly no server access
                 # As requested: show author info dialog only in this case
                 self.show_author_info()
            return

        if not info:
            self.update_info = None
            self.update_version_label(status="latest")
            if manual:
                QMessageBox.information(self, "Check for Updates", "You're already on the latest version.")
            return

        self.update_info = info
        self.update_version_label(status="update", latest_ver=info.get("version"))

        if show_prompt:
            self.show_update_prompt(info)

    def update_version_label(self, status="unknown", latest_ver=None):
        base = f"v{APP_VERSION} Desktop"
        text = base

        if status == "latest":
            text = f"{base} • Latest"
            self.version_label.setProperty("latest", True)
            self.version_label.setProperty("updateAvailable", False)
            self.version_label.setToolTip("You are on the latest version.")
            self.version_label.setCursor(Qt.CursorShape.ArrowCursor)
        elif status == "update":
            suffix = f"Update {latest_ver}" if latest_ver else "Update available"
            text = f"{base} • {suffix}"
            self.version_label.setProperty("latest", False)
            self.version_label.setProperty("updateAvailable", True)
            self.version_label.setToolTip("Click to update.")
            self.version_label.setCursor(Qt.CursorShape.PointingHandCursor)
        else:
            self.version_label.setProperty("latest", False)
            self.version_label.setProperty("updateAvailable", False)
            self.version_label.setToolTip("Version information.")
            self.version_label.setCursor(Qt.CursorShape.ArrowCursor)

        self.version_label.setText(text)
        self.version_label.style().unpolish(self.version_label)
        self.version_label.style().polish(self.version_label)

    def on_version_label_clicked(self):
        if self.update_info:
            self.show_update_prompt(self.update_info)
        else:
            self.start_update_check(show_prompt=True, manual=True)

    def show_update_prompt(self, info):
        manifest = info.get("manifest", {})
        latest_ver = info.get("version")
        notes = (manifest.get("notes") or "No release notes.").strip()
        release_dir = info.get("release_dir")

        msg = (
            f"New version {latest_ver} is available (current {CURRENT_VERSION}).\n\n"
            f"Notes:\n{notes}\n\n"
            "Update now?"
        )

        box = QMessageBox(self)
        box.setWindowTitle("Update Available")
        box.setText(msg)
        checkbox = QCheckBox("Show this update prompt on startup")
        checkbox.setChecked(get_update_prompt_on_startup())
        box.setCheckBox(checkbox)

        btn_update = box.addButton("Update", QMessageBox.ButtonRole.AcceptRole)
        btn_open = box.addButton("Open Folder", QMessageBox.ButtonRole.ActionRole)
        box.addButton(QMessageBox.StandardButton.Cancel)
        box.exec()

        set_update_prompt_on_startup(checkbox.isChecked())

        clicked = box.clickedButton()
        if clicked == btn_open:
            if release_dir and os.path.exists(release_dir):
                os.startfile(release_dir)
            else:
                QMessageBox.warning(self, "Open Folder", "Release folder not found.")
            return

        if clicked != btn_update:
            return

        self.start_update_from_info(info, notes)

    def start_update_from_info(self, info, notes):
        release_dir = info.get("release_dir")
        updater_path = info.get("updater_path")

        if not updater_path or not os.path.exists(updater_path):
            QMessageBox.information(
                self,
                "Update",
                "Updater not found. Opening release folder instead.",
            )
            if release_dir and os.path.exists(release_dir):
                os.startfile(release_dir)
            return

        if os.path.basename(sys.executable).lower().startswith("python"):
            QMessageBox.information(
                self,
                "Update",
                "You are running from source. Please run the updater from the release folder.",
            )
            if release_dir and os.path.exists(release_dir):
                os.startfile(release_dir)
            return

        payload = {
            "version": info.get("version"),
            "notes": notes,
            "session": self.collect_session_state(),
        }
        set_pending_update(payload)

        try:
            temp_updater = os.path.join(tempfile.gettempdir(), "DocCompareAI_Updater.exe")
            shutil.copy2(updater_path, temp_updater)
            install_dir = os.path.dirname(sys.executable)
            subprocess.Popen([temp_updater, release_dir, install_dir])
            QGuiApplication.instance().quit()
        except Exception as exc:
            set_pending_update({})
            QMessageBox.warning(self, "Update Failed", f"Could not start updater.\n\n{exc}")

    def collect_session_state(self):
        return {
            "file1_path": self.file1_path,
            "file2_path": self.file2_path,
            "sheet1_name": self.sheet1_name,
            "sheet2_name": self.sheet2_name,
            "result_view": self.stack.currentWidget() is self.result_view,
        }

    def restore_session(self, state):
        if not state:
            return

        missing = []
        file1 = state.get("file1_path")
        file2 = state.get("file2_path")

        if file1 and os.path.exists(file1):
            self.drop1.process_file(file1)
        elif file1:
            missing.append(file1)

        if file2 and os.path.exists(file2):
            self.drop2.process_file(file2)
        elif file2:
            missing.append(file2)

        s1 = state.get("sheet1_name")
        s2 = state.get("sheet2_name")

        if s1 and self.sheet1_combo.isEnabled():
            self.sheet1_combo.setCurrentText(s1)
        if s2 and self.sheet2_combo.isEnabled():
            self.sheet2_combo.setCurrentText(s2)

        if missing:
            QMessageBox.warning(
                self,
                "Restore Session",
                "Some files could not be found and were not restored:\n\n" + "\n".join(missing),
            )

        if state.get("result_view") and self.compare_btn.isEnabled():
            QTimer.singleShot(200, self.start_comparison)

    def handle_post_update(self):
        payload = pop_pending_update()
        if not payload:
            return

        session = payload.get("session") or {}
        notes = (payload.get("notes") or "").strip()
        version = payload.get("version") or APP_VERSION

        def _apply():
            if session:
                self.restore_session(session)
            if notes:
                msg = f"Updated to version {version}.\n\nNotes:\n{notes}"
            else:
                msg = f"Updated to version {version}."
            QMessageBox.information(self, "Update Complete", msg)

        QTimer.singleShot(0, _apply)

    def save_feedback(self, dialog, text):
        content = (text or "").strip()
        if not content:
            QMessageBox.information(self, "Feedback", "Please enter your feedback before submitting.")
            return
        try:
            _local_ok, remote_ok = submit_feedback(content)
            if remote_ok:
                QMessageBox.information(self, "Feedback", "Thanks! Your feedback has been saved and synced.")
            else:
                QMessageBox.information(self, "Feedback", "Thanks! Your feedback has been saved.")
        except Exception:
            QMessageBox.information(self, "Feedback", "Thanks! Your feedback has been received.")
        dialog.accept()

    def load_svg_colored_icon(self, name, color_hex, size=24):
        path = self.icon_dir / f"{name}.svg"
        if not path.exists():
            return None
        try:
            svg_text = path.read_text(encoding="utf-8")
        except Exception:
            return None

        def replace_color(match):
            val = match.group(1)
            if val.lower() == "none":
                return match.group(0)
            return f'{match.group(0).split("=")[0]}="{color_hex}"'

        svg_text = re.sub(r'stroke="([^"]+)"', replace_color, svg_text)
        svg_text = re.sub(r'fill="([^"]+)"', replace_color, svg_text)
        svg_text = re.sub(r'stroke:#([0-9A-Fa-f]{3,6})', f'stroke:{color_hex}', svg_text)
        svg_text = re.sub(r'fill:#([0-9A-Fa-f]{3,6})', f'fill:{color_hex}', svg_text)

        renderer = QSvgRenderer(QByteArray(svg_text.encode("utf-8")))
        pixmap = QPixmap(size, size)
        pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pixmap)
        renderer.render(painter)
        painter.end()
        return QIcon(pixmap)

    def setup_toggle_icons(self):
        self._toggle_icons.clear()

        def register(button, icon_name, fallback):
            normal = button.icon()
            checked = self.load_svg_colored_icon(icon_name, "#FFFFFF")
            if checked is None:
                checked = normal
            self._toggle_icons[button] = (normal, checked)
            self.apply_toggle_icon(button, button.isChecked())
            button.toggled.connect(lambda v, b=button: self.apply_toggle_icon(b, v))

        register(self.ribbon_diff_btn, "diff-toggle", QStyle.StandardPixmap.SP_DialogApplyButton)
        register(self.ribbon_diff_content_btn, "diff-content", QStyle.StandardPixmap.SP_FileDialogContentsView)
        register(self.ribbon_diff_format_btn, "diff-format", QStyle.StandardPixmap.SP_FileDialogInfoView)
        register(self.ribbon_diff_formula_btn, "diff-formula", QStyle.StandardPixmap.SP_FileDialogInfoView)
        register(self.ribbon_only_changes_btn, "diff-filter", QStyle.StandardPixmap.SP_DirIcon)

    def apply_toggle_icon(self, button, checked):
        pair = self._toggle_icons.get(button)
        if not pair:
            return
        normal, checked_icon = pair
        button.setIcon(checked_icon if checked else normal)

    def on_stack_changed(self, _index):
        is_result = self.stack.currentWidget() is self.result_view
        self.header.setVisible(not is_result)
        self.ribbon_home_panel.setEnabled(is_result)
        if hasattr(self, 'ribbon'):
            self.ribbon.setVisible(is_result)
        self.apply_ribbon_mode()

    def update_ribbon_sizes(self):
        if not hasattr(self, "ribbon_tabs"):
            return
        tab_h = self.ribbon_tabs.tabBar().sizeHint().height()
        panel_h = 0
        if hasattr(self, "ribbon_home_panel"):
            panel_h = max(panel_h, self.ribbon_home_panel.sizeHint().height())
        if hasattr(self, "ribbon_view_tab"):
            panel_h = max(panel_h, self.ribbon_view_tab.sizeHint().height())
        if hasattr(self, "ribbon_help_tab"):
            panel_h = max(panel_h, self.ribbon_help_tab.sizeHint().height())
        panel_h = max(panel_h, 58)  # Minimum panel height
        if hasattr(self, "ribbon_view_tab"):
            self.ribbon_view_tab.setMinimumHeight(panel_h)
        if hasattr(self, "ribbon_help_tab"):
            self.ribbon_help_tab.setMinimumHeight(panel_h)
        self.ribbon_full_height = tab_h + panel_h +0 # Padding
        self.ribbon_tabs.setMinimumHeight(self.ribbon_full_height)
        self.ribbon_tabs.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        if hasattr(self, "ribbon_home_scroll"):
            self.ribbon_home_scroll.setMinimumHeight(panel_h)

    def apply_ribbon_mode(self):
        if self.ribbon_mode == "groups":
            show_panel = True
            icon = QStyle.StandardPixmap.SP_TitleBarShadeButton
            tip = "Ribbon: Groups"
        else:
            show_panel = False
            icon = QStyle.StandardPixmap.SP_TitleBarMinButton
            tip = "Ribbon: Hidden"

        if hasattr(self, "ribbon_home_scroll"):
            self.ribbon_home_scroll.setVisible(show_panel)
        else:
            self.ribbon_home_panel.setVisible(show_panel)

        if show_panel:
            if hasattr(self, "ribbon_full_height"):
                self.ribbon_tabs.setMinimumHeight(self.ribbon_full_height)
            else:
                self.ribbon_tabs.setMinimumHeight(0)
            self.ribbon_tabs.setMaximumHeight(16777215)
        else:
            tab_h = self.ribbon_tabs.tabBar().sizeHint().height()
            target = max(24, tab_h + 6)
            self.ribbon_tabs.setMaximumHeight(target)
            self.ribbon_tabs.setMinimumHeight(target)

        self.ribbon_pin_btn.setIcon(self.style().standardIcon(icon))
        self.ribbon_pin_btn.setToolTip(tip)

    def on_result_view_mode_changed(self, mode):
        if not hasattr(self, "ribbon_view_group"):
            return
        self.ribbon_view_group.blockSignals(True)
        if mode == "page_break":
            self.ribbon_view_page_btn.setChecked(True)
        else:
            self.ribbon_view_normal_btn.setChecked(True)
        self.ribbon_view_group.blockSignals(False)

    def update_last_dir(self, path):
        if not path:
            return
        last_dir = os.path.dirname(path)
        if last_dir:
            self.last_open_dir = last_dir
            update_last_open_dir(last_dir)
            self.drop1.set_initial_dir(last_dir)
            self.drop2.set_initial_dir(last_dir)

    def load_sheet_names(self, file_path):
        if not file_path:
            return []
        try:
            wb = load_workbook(file_path, read_only=True)
            names = list(wb.sheetnames)
            wb.close()
            return names
        except Exception:
            return []

    def populate_sheet_combo(self, combo, file_path, is_left):
        previous = combo.currentText()
        names = self.load_sheet_names(file_path)
        if is_left:
            self.sheet1_names = names
        else:
            self.sheet2_names = names
        combo.blockSignals(True)
        combo.clear()
        if not names:
            combo.addItem("No sheets found")
            combo.setEnabled(False)
            if is_left:
                self.sheet1_name = None
            else:
                self.sheet2_name = None
            combo.blockSignals(False)
            self.check_ready()
            return

        combo.addItems(names)
        combo.setEnabled(True)
        if previous in names:
            combo.setCurrentText(previous)
        else:
            combo.setCurrentIndex(0)
        combo.blockSignals(False)

        if is_left:
            self.sheet1_name = combo.currentText()
        else:
            self.sheet2_name = combo.currentText()

    def on_sheet1_changed(self, name):
        self.sheet1_name = name if name and name != "No sheets found" else None
        self.check_ready()

    def on_sheet2_changed(self, name):
        self.sheet2_name = name if name and name != "No sheets found" else None
        self.check_ready()

    def on_result_sheet_change(self, side, sheet_name):
        if not sheet_name or sheet_name == "No sheets found":
            return
        if side == 1:
            if sheet_name == self.sheet1_name:
                return
            self.sheet1_name = sheet_name
            if self.sheet1_combo.isEnabled():
                self.sheet1_combo.setCurrentText(sheet_name)
        else:
            if sheet_name == self.sheet2_name:
                return
            self.sheet2_name = sheet_name
            if self.sheet2_combo.isEnabled():
                self.sheet2_combo.setCurrentText(sheet_name)
        # Trigger comparison again if ready
        if self.file1_path and self.file2_path and self.sheet1_name and self.sheet2_name:
            self.start_comparison()

    def apply_global_styles(self):
        self.setStyleSheet("""
            QWidget#AppRoot {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fafc, stop:0.5 #f1f5f9, stop:1 #e2e8f0);
            }
            QFrame#Ribbon {
                background-color: #ffffff;
                border-bottom: 1px solid #e2e8f0;
            }
            QWidget#RibbonPanel {
                background-color: transparent;
                border: none;
            }
            QScrollArea#RibbonScroll,
            QScrollArea#RibbonScroll > QWidget,
            QScrollArea#RibbonScroll > QWidget > QWidget {
                background: transparent;
                border: none;
            }
            QTabWidget#RibbonTabs::pane {
                border: none;
            }
            QTabWidget#RibbonTabs::pane {
                margin-top: 0px;
                padding-top: 0px;
            }
            QTabBar::tab {
                background: #f8fafc;
                color: #475569;
                padding: 4px 12px;
                border: 1px solid #e2e8f0;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                margin-right: 3px;
                margin-bottom: 0px;
            }
            QTabBar::tab:selected {
                background: #ffffff;
                color: #0f172a;
                font-weight: 700;
            }
            QFrame#RibbonGroup {
                background-color: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
            }
            QFrame#RibbonSeparator {
                color: #e2e8f0;
                min-width: 1px;
                max-width: 1px;
                margin: 10px 6px;
            }
            QToolButton#RibbonButton {
                background-color: transparent;
                border: none;
                color: #475569;
                padding: 4px 6px;
                border-radius: 6px;
                font-weight: 600;
                min-width: 62px;
                min-height: 46px;
            }
            QToolButton#RibbonButtonSmall {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                color: #475569;
                padding: 3px 6px;
                border-radius: 8px;
                font-weight: 600;
                min-width: 86px;
                min-height: 26px;
            }
            QToolButton#RibbonButtonSmall:hover {
                background-color: #e2e8f0;
            }
            QToolButton#RibbonButtonSmall:checked {
                background-color: #1e3a8a;
                color: #ffffff;
                border-color: #1e3a8a;
            }
            QToolButton#RibbonButton:hover {
                background-color: #e2e8f0;
            }
            QToolButton#RibbonButton:checked {
                background-color: #1e3a8a;
                color: #ffffff;
                border-color: #1e3a8a;
            }
            QLabel#RibbonGroupTitle {
                font-size: 10px;
                color: #64748b;
                font-weight: 700;
            }
            QToolButton#RibbonPin {
                background-color: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 2px 6px;
                min-width: 28px;
                min-height: 24px;
            }
            QToolButton#RibbonPin:hover {
                background-color: #e2e8f0;
            }
            QFrame#TopBar {
                background-color: #ffffff;
                border-bottom: 1px solid #e2e8f0;
            }
            QLabel#AppTitle {
                font-size: 20px;
                font-weight: 700;
                color: #0f172a;
            }
            QLabel#AppTagline {
                font-size: 11px;
                color: #64748b;
            }
            QFrame#ModePill {
                background-color: #f1f5f9;
                border: 1px solid #e2e8f0;
                border-radius: 16px;
            }
            QPushButton[modeTab="true"] {
                background-color: transparent;
                color: #475569;
                border: none;
                padding: 6px 14px;
                border-radius: 12px;
                font-weight: 600;
            }
            QPushButton[modeTab="true"]:checked {
                background-color: #0f172a;
                color: #ffffff;
            }
            QPushButton[modeTab="true"]:hover {
                background-color: #e2e8f0;
            }
            QLabel#VersionLabel {
                font-size: 11px;
                color: #94a3b8;
            }
            QLabel#VersionLabel[latest="true"] {
                color: #16a34a;
                font-weight: 700;
            }
            QLabel#VersionLabel[updateAvailable="true"] {
                color: #dc2626;
                font-weight: 700;
            }
            QFrame#FooterBar {
                background: transparent;
            }
            QLabel#LogoLabel {
                min-width: 120px;
                min-height: 40px;
                background-color: #ffffff;
                padding: 0px;
                margin: 0px;
                border: none;
            }
            QLabel#LogoLabel[placeholder="true"] {
                border: 1px dashed #cbd5e1;
                color: #94a3b8;
                padding: 4px 8px;
                font-size: 11px;
            }
            QLabel#PageTitle {
                font-size: 32px;
                font-weight: 700;
                color: #0f172a;
            }
            QLabel#PageSubtitle {
                font-size: 14px;
                color: #475569;
            }
            QLabel[chip="true"] {
                background-color: #e2e8f0;
                color: #334155;
                padding: 4px 10px;
                border-radius: 10px;
                font-size: 11px;
                font-weight: 600;
            }
            QFrame#UploadCard {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 16px;
            }
            QLabel#CardTitle {
                font-size: 16px;
                font-weight: 700;
                color: #0f172a;
            }
            QLabel#CardSubtitle {
                font-size: 12px;
                color: #64748b;
            }
            QLabel#HelperText {
                font-size: 11px;
                color: #94a3b8;
            }
            QLineEdit#NameBox {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 4px 8px;
                color: #334155;
                font-size: 12px;
            }
            QLineEdit#FormulaBar {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 4px 8px;
                color: #334155;
                font-size: 12px;
            }
            QLabel#SheetLabel {
                font-size: 11px;
                color: #64748b;
                font-weight: 700;
            }
            QToolButton#SheetButton {
                background-color: transparent;
                border: none;
                color: #0f172a;
                padding: 2px 6px;
                font-size: 12px;
                font-weight: 700;
            }
            QToolButton#SheetButton:hover {
                background-color: #e2e8f0;
                border-radius: 6px;
            }
            QComboBox#SheetCombo {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 4px 8px;
                min-height: 26px;
                color: #334155;
                font-size: 12px;
            }
            QComboBox#SheetCombo::drop-down {
                border: none;
                width: 18px;
            }
            QPushButton#PrimaryButton {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #14b8a6, stop:1 #6366f1);
                color: #ffffff;
                font-size: 14px;
                font-weight: 700;
                padding: 10px 22px;
                border-radius: 12px;
            }
            QPushButton#PrimaryButton:hover {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #0f766e, stop:1 #4f46e5);
            }
            QPushButton#PrimaryButton:disabled {
                background-color: #e2e8f0;
                color: #94a3b8;
            }
            QPushButton#GhostButton {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                color: #475569;
                padding: 6px 12px;
                border-radius: 10px;
                font-weight: 600;
            }
            QPushButton#GhostButton:hover {
                background-color: #f1f5f9;
            }
            QPushButton#ToolbarButton {
                background-color: #f1f5f9;
                border: 1px solid #e2e8f0;
                color: #475569;
                padding: 6px 10px;
                border-radius: 10px;
                font-weight: 600;
            }
            QPushButton#ToolbarButton:hover {
                background-color: #e2e8f0;
            }
            QPushButton#ToolbarButton:checked {
                background-color: #0f172a;
                color: #ffffff;
                border-color: #0f172a;
            }
            QLabel#ResultTitle {
                font-size: 16px;
                font-weight: 700;
                color: #0f172a;
            }
            QLabel[panelTitle="true"] {
                font-size: 12px;
                font-weight: 700;
                color: #475569;
                padding: 4px 0;
            }
            QLabel#ComingSoonTitle {
                font-size: 26px;
                font-weight: 700;
                color: #94a3b8;
            }
            QLabel#ComingSoonStatus {
                font-size: 14px;
                color: #cbd5f5;
            }
            QFrame#ResultStatusBar {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                min-height: 28px;
            }
            QLabel#StatusInfo {
                font-size: 10px;
                color: #475569;
                font-weight: 600;
            }
            QToolButton#StatusToggle {
                background-color: transparent;
                border: none;
                color: #475569;
                padding: 2px;
                border-radius: 6px;
                min-width: 24px;
                min-height: 24px;
            }
            QToolButton#StatusToggle:checked {
                background-color: #e2e8f0;
                color: #0f172a;
            }
            QToolButton#StatusZoomButton {
                background-color: #f8fafc;
                border: 1px solid #e2e8f0;
                color: #475569;
                border-radius: 6px;
                min-width: 20px;
                min-height: 20px;
                font-weight: 700;
            }
            QSlider#StatusZoomSlider::groove:horizontal {
                height: 4px;
                background: #e2e8f0;
                border-radius: 3px;
            }
            QSlider#StatusZoomSlider::handle:horizontal {
                background: #0f172a;
                width: 10px;
                margin: -3px 0;
                border-radius: 5px;
            }
            QLabel#StatusZoomLabel {
                font-size: 10px;
                color: #475569;
                font-weight: 700;
                min-width: 44px;
            }
            QFrame#ResultSide {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
            }
            QSplitter::handle {
                background: #e2e8f0;
            }
            QSplitter::handle:horizontal {
                width: 6px;
                margin: 0 4px;
            }
            QProgressBar#StatusProgress {
                border: 1px solid #e2e8f0;
                background: #f8fafc;
                border-radius: 4px;
                height: 8px;
            }
            QProgressBar#StatusProgress::chunk {
                background-color: #0f172a;
                border-radius: 4px;
            }
        """)
        self.update_ribbon_sizes()
