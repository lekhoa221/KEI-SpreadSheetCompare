import pandas as pd
import numpy as np


from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.writer.theme import theme_xml
import xml.etree.ElementTree as ET

THEME_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}

def _normalize_hex(value):
    if not value:
        return None
    value = str(value).strip().upper()
    if len(value) == 8:
        # ARGB
        if value == "00000000":
            return None
        value = value[2:]
    if len(value) == 6:
        return f"#{value}"
    return None

def _apply_tint(hex_color, tint):
    if not hex_color or tint is None or tint == 0:
        return hex_color
    tint = float(tint)
    if not hex_color.startswith("#") or len(hex_color) != 7:
        return hex_color
    r = int(hex_color[1:3], 16)
    g = int(hex_color[3:5], 16)
    b = int(hex_color[5:7], 16)
    def adj(c):
        if tint < 0:
            return int(c * (1 + tint))
        return int(c + (255 - c) * tint)
    r, g, b = adj(r), adj(g), adj(b)
    r = max(0, min(255, r))
    g = max(0, min(255, g))
    b = max(0, min(255, b))
    return f"#{r:02X}{g:02X}{b:02X}"

def _get_theme_colors(wb):
    theme_data = getattr(wb, "loaded_theme", None) or theme_xml
    if isinstance(theme_data, bytes):
        theme_data = theme_data.decode("utf-8", errors="ignore")
    try:
        root = ET.fromstring(theme_data)
    except Exception:
        return []
    clr_scheme = root.find(".//a:clrScheme", THEME_NS)
    if clr_scheme is None:
        return []
    colors = []
    for child in clr_scheme:
        srgb = child.find("a:srgbClr", THEME_NS)
        if srgb is not None and srgb.attrib.get("val"):
            colors.append(f"#{srgb.attrib['val'].upper()}")
            continue
        sysclr = child.find("a:sysClr", THEME_NS)
        if sysclr is not None:
            val = sysclr.attrib.get("lastClr") or sysclr.attrib.get("val")
            colors.append(f"#{val.upper()}" if val else None)
        else:
            colors.append(None)
    return colors

def _safe_color_attr(color, attr):
    try:
        val = getattr(color, attr)
    except Exception:
        return None
    if isinstance(val, str) and val.startswith("Values must be of type"):
        return None
    return val

def _color_to_hex(color, wb, theme_colors, role=None):
    if not color:
        return None
    auto_val = _safe_color_attr(color, "auto")
    if auto_val is True:
        return None
    color_type = _safe_color_attr(color, "type")
    if color_type == "auto":
        return None
    if color_type == "rgb":
        rgb_val = _safe_color_attr(color, "rgb")
        if rgb_val:
            hex_rgb = _normalize_hex(rgb_val)
            if hex_rgb:
                return hex_rgb
    indexed_val = _safe_color_attr(color, "indexed")
    if indexed_val is not None:
        try:
            idx = int(indexed_val)
        except Exception:
            idx = None
        if idx is not None and idx >= 0 and idx < len(COLOR_INDEX):
            return _normalize_hex(COLOR_INDEX[idx])
    theme_val = _safe_color_attr(color, "theme")
    if color_type == "theme" or theme_val is not None:
        try:
            theme_idx = int(theme_val)
        except Exception:
            theme_idx = None
        if theme_idx is not None:
            # Some files use theme=1 for default font color (should be dark text).
            if role == "font" and theme_idx == 1 and len(theme_colors) > 1:
                if str(theme_colors[1]).upper() in ("#FFFFFF", "FFFFFF"):
                    base = theme_colors[0]
                    tint_val = _safe_color_attr(color, "tint")
                    return _apply_tint(base, tint_val)
            if theme_idx >= 0 and theme_idx < len(theme_colors):
                base = theme_colors[theme_idx]
                tint_val = _safe_color_attr(color, "tint")
                return _apply_tint(base, tint_val)
    return None

def _fill_color_to_hex(fill, wb, theme_colors):
    if not fill:
        return None
    candidates = []
    if hasattr(fill, "fgColor"):
        candidates.append(fill.fgColor)
    if hasattr(fill, "start_color"):
        candidates.append(fill.start_color)
    if hasattr(fill, "bgColor"):
        candidates.append(fill.bgColor)
    if hasattr(fill, "end_color"):
        candidates.append(fill.end_color)

    for color in candidates:
        hex_code = _color_to_hex(color, wb, theme_colors, role="fill")
        if hex_code:
            return hex_code
    return None

def load_excel_with_styles(file_path, sheet_name=None):
    """
    Load an Excel file including its styles and merged cells.
    Returns: (DataFrame, styles_dict, merges_list)
    """
    try:
        # data_only=True gets values instead of formulas
        wb_values = load_workbook(file_path, data_only=True)
        wb_formulas = load_workbook(file_path, data_only=False)
        if sheet_name:
            if sheet_name not in wb_values.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
            ws = wb_values[sheet_name]
            ws_formulas = wb_formulas[sheet_name]
        else:
            ws = wb_values.active # Default to first sheet
            ws_formulas = wb_formulas.active
        theme_colors = _get_theme_colors(wb_values)
        
        data = []
        styles = {}
        formulas = {}
        merges = []
        
        # 1. Extract Merged Cells
        # Openpyxl ranges are 1-based indices
        for rng in ws.merged_cells.ranges:
            # Format: (row_idx, col_idx, row_span, col_span) - 0-based for Qt
            r_idx = rng.min_row - 1
            c_idx = rng.min_col - 1
            r_span = rng.max_row - rng.min_row + 1
            c_span = rng.max_col - rng.min_col + 1
            merges.append((r_idx, c_idx, r_span, c_span))

        # 2. Iterate Rows for Data and Styles
        # We assume a rectangular grid based on max dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        for r_idx, row in enumerate(ws.iter_rows(max_row=max_row, max_col=max_col)):
            row_data = []
            for c_idx, cell in enumerate(row):
                # Value
                row_data.append(cell.value)

                # Formula
                try:
                    formula_cell = ws_formulas.cell(row=r_idx + 1, column=c_idx + 1)
                    if formula_cell.data_type == 'f' or (isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")):
                        formulas[(r_idx, c_idx)] = formula_cell.value
                except Exception:
                    pass
                
                # Style Extraction
                cell_style = {}
                
                # Background Color
                # Handling openpyxl color complexities slightly better
                # We skip '00000000' if it appears as it often usually means transparent/auto in some contexts, but 'FF000000' is black.
                if cell.fill:
                    hex_code = _fill_color_to_hex(cell.fill, wb_values, theme_colors)
                    if hex_code:
                        cell_style['bg'] = hex_code

                # Font
                if cell.font:
                    font_style = {
                        'bold': cell.font.bold,
                        'italic': cell.font.italic,
                        'name': cell.font.name,
                        'size': cell.font.size
                    }
                    if cell.font.color:
                        font_color = _color_to_hex(cell.font.color, wb_values, theme_colors, role="font")
                        if font_color:
                            font_style['color'] = font_color
                    
                    cell_style['font'] = font_style

                # Alignment
                if cell.alignment:
                    cell_style['align'] = {
                        'horizontal': cell.alignment.horizontal,
                        'vertical': cell.alignment.vertical,
                        'wrap_text': cell.alignment.wrap_text
                    }

                # Number Format
                if cell.number_format:
                    cell_style['numfmt'] = cell.number_format
                
                # Borders
                # Qt TableWidget doesn't support individual borders easily without a Delegate.
                # We will extract it, but might not render it fully yet.
                if cell.border:
                    # Just checking if any border side exists
                    border_style = {}
                    if cell.border.left and cell.border.left.style: border_style['left'] = True
                    if cell.border.right and cell.border.right.style: border_style['right'] = True
                    if cell.border.top and cell.border.top.style: border_style['top'] = True
                    if cell.border.bottom and cell.border.bottom.style: border_style['bottom'] = True
                    
                    if border_style:
                        cell_style['border'] = border_style

                if cell_style:
                    styles[(r_idx, c_idx)] = cell_style

            data.append(row_data)

        # 3. Extract Column Widths and Row Heights
        col_widths = {}
        row_heights = {}
        
        # Column Widths
        # OpenPyXL column dimensions are 1-indexed (A, B, C...)
        from openpyxl.utils.cell import column_index_from_string
        for col_char, dim in ws.column_dimensions.items():
            if dim.width:
                col_idx = column_index_from_string(col_char) - 1
                col_widths[col_idx] = dim.width
            
        # Row Heights
        for row_idx, dim in ws.row_dimensions.items():
            if dim.height:
                row_heights[row_idx - 1] = dim.height

        df = pd.DataFrame(data)
        return df, styles, merges, formulas, col_widths, row_heights

    except Exception as e:
        raise ValueError(f"Error loading file with styles: {str(e)}")

def load_excel(file_path):
    # Backward compatibility wrapper or just use pandas for pure data
    return load_excel_with_styles(file_path)[0]


def compare_dataframes(df_old: pd.DataFrame, df_new: pd.DataFrame):
    """
    Compare two DataFrames and return a summary of changes.
    """
    # 1. Align Columns (by index since we don't have headers)
    max_cols = max(df_old.shape[1], df_new.shape[1])
    
    # Reindex columns to ensure matching shape
    df_old = df_old.reindex(columns=range(max_cols))
    df_new = df_new.reindex(columns=range(max_cols))
    
    # 2. Align Rows
    max_rows = max(len(df_old), len(df_new))
    df_old = df_old.reindex(index=range(max_rows))
    df_new = df_new.reindex(index=range(max_rows))

    changes = []
    
    # 3. Iterate and Compare cell-by-cell
    # Fill NaN with a placeholder
    df_old_filled = df_old.fillna("###NULL###")
    df_new_filled = df_new.fillna("###NULL###")
    
    diff_mask = (df_old_filled != df_new_filled)
    
    # Get coordinates of differences
    diff_locations = np.where(diff_mask)
    rows = diff_locations[0]
    cols = diff_locations[1]
    
    for r, c in zip(rows, cols):
        old_val = df_old.iloc[r, c]
        new_val = df_new.iloc[r, c]
        
        changes.append({
            "row": int(r),
            "col": int(c),
            "old": str(old_val) if pd.notna(old_val) else "",
            "new": str(new_val) if pd.notna(new_val) else "",
            "type": "modified"
        })

    # Summary Stats
    summary = {
        "total_rows": int(max_rows),
        "total_cols": int(max_cols),
        "changes_count": len(changes)
    }

    return {
        "changes": changes,
        "summary": summary
    }
